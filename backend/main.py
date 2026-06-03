from datetime import datetime
from io import BytesIO
import os

from fastapi import Depends, FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from fastapi.security import OAuth2PasswordRequestForm
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook
from sqlalchemy import or_
from sqlalchemy.orm import Session

from .auth import (
    ROLE_MEMBER,
    ROLE_OWNER,
    ROLE_READONLY,
    authenticate_user,
    create_access_token,
    get_current_user,
    get_password_hash,
    require_roles,
)
from .database import Base, SessionLocal, engine, get_db
from .models import Consumable, Equipment, StockRecord, User
from .schemas import (
    ConsumableIn,
    ConsumableOut,
    EquipmentIn,
    EquipmentOut,
    RecordOut,
    SearchOut,
    StockAction,
    Token,
    UserCreate,
    UserOut,
    UserUpdate,
)

Base.metadata.create_all(bind=engine)

app = FastAPI(title="实验室资产与耗材共享台账系统")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://127.0.0.1:5173"],
    allow_origin_regex=r"http://(localhost|127\.0\.0\.1):\d+",
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def seed_admin():
    db = SessionLocal()
    try:
        if not db.query(User).filter(User.username == "admin").first():
            admin_password = os.getenv("LAB_ADMIN_PASSWORD", "admin123")
            db.add(
                User(
                    username="admin",
                    full_name="实验室负责人",
                    role=ROLE_OWNER,
                    hashed_password=get_password_hash(admin_password),
                    is_active=True,
                )
            )
            db.commit()
    finally:
        db.close()


seed_admin()

app.mount("/app", StaticFiles(directory="static", html=True), name="app")


def touch(obj):
    obj.updated_at = datetime.utcnow()
    return obj


@app.post("/api/auth/login", response_model=Token)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    user = authenticate_user(db, form_data.username, form_data.password)
    if not user:
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": user.username, "role": user.role})
    return {"access_token": token, "token_type": "bearer", "user": user}


@app.get("/api/auth/me", response_model=UserOut)
def me(current_user: User = Depends(get_current_user)):
    return current_user


@app.get("/api/search", response_model=SearchOut)
def search(q: str = "", db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    like = f"%{q.strip()}%"
    consumables = db.query(Consumable)
    equipment = db.query(Equipment)
    if q.strip():
        consumables = consumables.filter(
            or_(
                Consumable.name.like(like),
                Consumable.category.like(like),
                Consumable.specification.like(like),
                Consumable.location.like(like),
            )
        )
        equipment = equipment.filter(
            or_(
                Equipment.asset_no.like(like),
                Equipment.name.like(like),
                Equipment.model.like(like),
                Equipment.location.like(like),
                Equipment.owner.like(like),
            )
        )
    return {"consumables": consumables.limit(30).all(), "equipment": equipment.limit(30).all()}


@app.get("/api/consumables", response_model=list[ConsumableOut])
def list_consumables(q: str = "", db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    query = db.query(Consumable)
    if q.strip():
        like = f"%{q.strip()}%"
        query = query.filter(or_(Consumable.name.like(like), Consumable.category.like(like), Consumable.location.like(like)))
    return query.order_by(Consumable.updated_at.desc()).all()


@app.post("/api/consumables", response_model=ConsumableOut)
def create_consumable(payload: ConsumableIn, db: Session = Depends(get_db), _: User = Depends(require_roles([ROLE_OWNER, ROLE_MEMBER]))):
    item = Consumable(**payload.dict())
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


@app.put("/api/consumables/{item_id}", response_model=ConsumableOut)
def update_consumable(item_id: int, payload: ConsumableIn, db: Session = Depends(get_db), _: User = Depends(require_roles([ROLE_OWNER, ROLE_MEMBER]))):
    item = db.get(Consumable, item_id)
    if not item:
        raise HTTPException(status_code=404, detail="耗材不存在")
    for key, value in payload.dict().items():
        setattr(item, key, value)
    touch(item)
    db.commit()
    db.refresh(item)
    return item


@app.delete("/api/consumables/{item_id}")
def delete_consumable(item_id: int, db: Session = Depends(get_db), _: User = Depends(require_roles([ROLE_OWNER]))):
    item = db.get(Consumable, item_id)
    if not item:
        raise HTTPException(status_code=404, detail="耗材不存在")
    db.delete(item)
    db.commit()
    return {"ok": True}


@app.get("/api/equipment", response_model=list[EquipmentOut])
def list_equipment(q: str = "", db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    query = db.query(Equipment)
    if q.strip():
        like = f"%{q.strip()}%"
        query = query.filter(or_(Equipment.name.like(like), Equipment.asset_no.like(like), Equipment.location.like(like)))
    return query.order_by(Equipment.updated_at.desc()).all()


@app.post("/api/equipment", response_model=EquipmentOut)
def create_equipment(payload: EquipmentIn, db: Session = Depends(get_db), _: User = Depends(require_roles([ROLE_OWNER, ROLE_MEMBER]))):
    if db.query(Equipment).filter(Equipment.asset_no == payload.asset_no).first():
        raise HTTPException(status_code=409, detail="资产编号已存在")
    item = Equipment(**payload.dict())
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


@app.put("/api/equipment/{item_id}", response_model=EquipmentOut)
def update_equipment(item_id: int, payload: EquipmentIn, db: Session = Depends(get_db), _: User = Depends(require_roles([ROLE_OWNER, ROLE_MEMBER]))):
    item = db.get(Equipment, item_id)
    if not item:
        raise HTTPException(status_code=404, detail="设备不存在")
    duplicate = db.query(Equipment).filter(Equipment.asset_no == payload.asset_no, Equipment.id != item_id).first()
    if duplicate:
        raise HTTPException(status_code=409, detail="资产编号已存在")
    for key, value in payload.dict().items():
        setattr(item, key, value)
    touch(item)
    db.commit()
    db.refresh(item)
    return item


@app.delete("/api/equipment/{item_id}")
def delete_equipment(item_id: int, db: Session = Depends(get_db), _: User = Depends(require_roles([ROLE_OWNER]))):
    item = db.get(Equipment, item_id)
    if not item:
        raise HTTPException(status_code=404, detail="设备不存在")
    db.delete(item)
    db.commit()
    return {"ok": True}


@app.post("/api/stock/inbound", response_model=RecordOut)
def inbound(payload: StockAction, db: Session = Depends(get_db), current_user: User = Depends(require_roles([ROLE_OWNER, ROLE_MEMBER]))):
    if payload.item_type != "consumable":
        raise HTTPException(status_code=400, detail="入库当前仅支持耗材")
    item = db.get(Consumable, payload.item_id)
    if not item:
        raise HTTPException(status_code=404, detail="耗材不存在")
    item.quantity += payload.quantity
    touch(item)
    record = StockRecord(**payload.dict(), action="inbound", operator_id=current_user.id)
    db.add(record)
    db.commit()
    db.refresh(record)
    return record


@app.post("/api/stock/checkout", response_model=RecordOut)
def checkout(payload: StockAction, db: Session = Depends(get_db), current_user: User = Depends(require_roles([ROLE_OWNER, ROLE_MEMBER]))):
    if payload.item_type != "consumable":
        raise HTTPException(status_code=400, detail="领用当前仅支持耗材")
    item = db.get(Consumable, payload.item_id)
    if not item:
        raise HTTPException(status_code=404, detail="耗材不存在")
    if item.quantity < payload.quantity:
        raise HTTPException(status_code=400, detail="库存不足")
    item.quantity -= payload.quantity
    touch(item)
    record = StockRecord(**payload.dict(), action="checkout", operator_id=current_user.id)
    db.add(record)
    db.commit()
    db.refresh(record)
    return record


@app.get("/api/records", response_model=list[RecordOut])
def records(action: str = "", db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    query = db.query(StockRecord)
    if action:
        query = query.filter(StockRecord.action == action)
    return query.order_by(StockRecord.created_at.desc()).limit(300).all()


@app.get("/api/users", response_model=list[UserOut])
def list_users(db: Session = Depends(get_db), _: User = Depends(require_roles([ROLE_OWNER]))):
    return db.query(User).order_by(User.created_at.desc()).all()


@app.post("/api/users", response_model=UserOut)
def create_user(payload: UserCreate, db: Session = Depends(get_db), _: User = Depends(require_roles([ROLE_OWNER]))):
    if payload.role not in {ROLE_OWNER, ROLE_MEMBER, ROLE_READONLY}:
        raise HTTPException(status_code=400, detail="角色无效")
    if db.query(User).filter(User.username == payload.username).first():
        raise HTTPException(status_code=409, detail="用户名已存在")
    user = User(
        username=payload.username,
        full_name=payload.full_name,
        role=payload.role,
        hashed_password=get_password_hash(payload.password),
        is_active=payload.is_active,
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    return user


@app.put("/api/users/{user_id}", response_model=UserOut)
def update_user(user_id: int, payload: UserUpdate, db: Session = Depends(get_db), _: User = Depends(require_roles([ROLE_OWNER]))):
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")
    if payload.role not in {ROLE_OWNER, ROLE_MEMBER, ROLE_READONLY}:
        raise HTTPException(status_code=400, detail="角色无效")
    user.full_name = payload.full_name
    user.role = payload.role
    user.is_active = payload.is_active
    if payload.password:
        user.hashed_password = get_password_hash(payload.password)
    db.commit()
    db.refresh(user)
    return user


@app.get("/api/export/excel")
def export_excel(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    wb = Workbook()
    ws = wb.active
    ws.title = "耗材库存"
    ws.append(["名称", "分类", "规格", "单位", "库存", "预警线", "位置", "供应商", "备注"])
    for item in db.query(Consumable).order_by(Consumable.name).all():
        ws.append([item.name, item.category, item.specification, item.unit, item.quantity, item.min_quantity, item.location, item.supplier, item.remark])

    ws = wb.create_sheet("设备台账")
    ws.append(["资产编号", "名称", "型号", "负责人", "状态", "位置", "购置日期", "价格", "备注"])
    for item in db.query(Equipment).order_by(Equipment.asset_no).all():
        ws.append([item.asset_no, item.name, item.model, item.owner, item.status, item.location, item.purchase_date, item.price, item.remark])

    ws = wb.create_sheet("历史记录")
    ws.append(["类型", "对象ID", "动作", "数量", "操作人", "领用人", "用途", "备注", "时间"])
    for item in db.query(StockRecord).order_by(StockRecord.created_at.desc()).all():
        ws.append([item.item_type, item.item_id, item.action, item.quantity, item.operator.full_name or item.operator.username, item.borrower, item.purpose, item.remark, item.created_at.strftime("%Y-%m-%d %H:%M:%S")])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    filename = f"lab-ledger-{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
